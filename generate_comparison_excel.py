#!/usr/bin/env python3
"""Generate Old vs New mapping comparison Excel (63 funds, 2 tabs: mean + last date).

Usage: python3 generate_comparison_excel.py \
         rolling_stats_OLD.csv rolling_stats_NEW.csv \
         rolling_params_OLD.csv rolling_params_NEW.csv \
         [mapping.csv] [output.xlsx]
"""
import sys, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

args = sys.argv[1:]
if len(args) < 4:
    print("Usage: generate_comparison_excel.py stats_OLD stats_NEW params_OLD params_NEW [mapping.csv] [output.xlsx]")
    sys.exit(1)

old_s = pd.read_csv(args[0], low_memory=False)
new_s = pd.read_csv(args[1], low_memory=False)
old_p = pd.read_csv(args[2], low_memory=False)
new_p = pd.read_csv(args[3], low_memory=False)
mapping_csv = args[4] if len(args) > 4 else None
OUTPUT = args[5] if len(args) > 5 else "Old_vs_New_Mapping_63_Funds.xlsx"

# Exclude last date artifact if present
for df in [old_s, new_s, old_p, new_p]:
    df.drop(df[df['as_of'] == '2026-06-26'].index, inplace=True, errors='ignore')

# Identify changed funds from mapping (if provided)
changed = set()
if mapping_csv:
    prod = pd.read_csv(mapping_csv)
    prod = prod[prod['nav_asset_id'].notna()]
    changed = set(prod[prod['fund_type'] == 'Private Equity']['label'].values)
    for _, r in prod.iterrows():
        if 'Value-Added' in r['label'] and 'Ex-' not in r['label']:
            changed.add(r['label'])

date = sorted(old_s['as_of'].unique())[-1]  # last date

# Styles
thin = Side(style='thin', color='D0D0D0'); brd = Border(top=thin, bottom=thin, left=thin, right=thin)
HDR = PatternFill('solid', fgColor='2F5496')
DIFF_HDR = PatternFill('solid', fgColor='1B3A5C')
CHG = PatternFill('solid', fgColor='FFF2CC')
hf = Font(name='Arial', size=8, bold=True, color='FFFFFF')
bf = Font(name='Arial', size=7); bfb = Font(name='Arial', size=7, bold=True)
rf = Font(name='Arial', size=7, color='C00000', bold=True)
ga = Alignment(horizontal='center', vertical='center', wrap_text=True)

headers = [
    ("Fund", 38, False), ("CCY", 4, False), ("Type", 14, False), ("Chg?", 5, False),
    ("Vol\nOLD", 9, False), ("Vol\nNEW", 9, False), ("Vol\nDiff", 7, True),
    ("OLS Beta\nOLD", 9, False), ("OLS Beta\nNEW", 9, False), ("OLS\nDiff", 7, True),
    ("Beta(r*,SPY)\nOLD", 11, False), ("Beta(r*,SPY)\nNEW", 11, False), ("SPY\nDiff", 7, True),
]

wb = Workbook(); wb.remove(wb.active)

for mode, tab_name in [('mean', f'Mean ({old_s["as_of"].nunique()} dates)'), ('last', f'Last Date ({date})')]:
    ws = wb.create_sheet(tab_name)

    if mode == 'mean':
        os_ = old_s.groupby('asset_id').mean(numeric_only=True)
        ns_ = new_s.groupby('asset_id').mean(numeric_only=True)
        op_ = old_p[old_p['beta_ols'].abs() > 0.01].groupby('ticker').mean(numeric_only=True)
        np_ = new_p[new_p['beta_ols'].abs() > 0.01].groupby('ticker').mean(numeric_only=True)
    else:
        os_ = old_s[old_s['as_of'] == date].set_index('asset_id')
        ns_ = new_s[new_s['as_of'] == date].set_index('asset_id')
        op_ = old_p[old_p['as_of'] == date].set_index('ticker')
        np_ = new_p[new_p['as_of'] == date].set_index('ticker')

    for i, (h, w, is_diff) in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hf; c.fill = DIFF_HDR if is_diff else HDR; c.alignment = ga; c.border = brd
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 28

    all_funds = sorted(os_.index)
    for ri, fund in enumerate(all_funds, 2):
        parts = fund.split(' | ')
        ccy = parts[0].strip()
        name = parts[1].strip().replace(f' - {ccy}', '') if len(parts) > 1 else ''
        is_chg = fund in changed

        def gv(df, col, idx):
            if idx in df.index and col in df.columns and pd.notna(df.loc[idx, col]):
                return float(df.loc[idx, col])
            return None

        ov = gv(os_, 'wk_ann_vol', fund); nv = gv(ns_, 'wk_ann_vol', fund)
        oo = gv(op_, 'beta_ols', fund);   no = gv(np_, 'beta_ols', fund)
        ob = gv(os_, 'beta_rstar_spy', fund); nb = gv(ns_, 'beta_rstar_spy', fund)

        ws.cell(row=ri, column=1, value=name).font = bf
        ws.cell(row=ri, column=2, value=ccy).font = bf; ws.cell(row=ri, column=2).alignment = ga
        ws.cell(row=ri, column=3).font = bf
        ws.cell(row=ri, column=4, value='YES' if is_chg else '').font = bfb; ws.cell(row=ri, column=4).alignment = ga
        if is_chg:
            for c in range(1, 14): ws.cell(row=ri, column=c).fill = CHG

        for col, val in [(5, ov), (6, nv)]:
            if val is not None:
                ws.cell(row=ri, column=col, value=round(val, 6)).font = bf
                ws.cell(row=ri, column=col).number_format = '0.00%'
        if ov is not None and nv is not None:
            d = nv - ov
            ws.cell(row=ri, column=7, value=round(d, 6)).font = rf if abs(d) > 0.003 else bf
            ws.cell(row=ri, column=7).number_format = '+0.00%;-0.00%'

        for col, val in [(8, oo), (9, no)]:
            if val is not None and abs(val) > 0.01:
                ws.cell(row=ri, column=col, value=round(val, 4)).font = bf
                ws.cell(row=ri, column=col).number_format = '0.000'
        if oo is not None and no is not None and abs(oo) > 0.01 and abs(no) > 0.01:
            d = no - oo
            ws.cell(row=ri, column=10, value=round(d, 4)).font = rf if abs(d) > 0.03 else bf
            ws.cell(row=ri, column=10).number_format = '+0.000;-0.000'

        for col, val in [(11, ob), (12, nb)]:
            if val is not None:
                ws.cell(row=ri, column=col, value=round(val, 4)).font = bf
                ws.cell(row=ri, column=col).number_format = '0.000'
        if ob is not None and nb is not None:
            d = nb - ob
            ws.cell(row=ri, column=13, value=round(d, 4)).font = rf if abs(d) > 0.03 else bf
            ws.cell(row=ri, column=13).number_format = '+0.000;-0.000'

        for c in range(1, 14): ws.cell(row=ri, column=c).border = brd

wb.save(OUTPUT)
print(f"Saved: {OUTPUT} ({len(all_funds)} funds, changed={len(changed)})")
