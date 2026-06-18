#!/usr/bin/env python3
"""Generate JPM Portfolio Risk Excel with ELRISKMEASURE/ELSTRESSSCENARIO formulas.

Usage: python3 generate_portfolio_risk.py [output.xlsx]

Creates 10 portfolios (7 public + 1 PI at 20% weight) with:
  - Vol, ES, VaR: positions + contributions (array formulas per position row)
  - Portfolio totals: single formula in the TOTAL row
  - Stress: 2008 GFC (3021), USD +100bps (21148), USD -100bps (21157)
"""
import sys
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = sys.argv[1] if len(sys.argv) > 1 else "JPM_Portfolio_Risk_10_Funds.xlsx"

PI_FUNDS = [
    ("KKR North America XIV",       "8104697100", "04cc4f53-e24f-45aa-85d0-5e975e67bfbd", "Core PE / Large Cap"),
    ("Riverstone IV",                "8104697100", "04cc4f53-e24f-45aa-85d0-5e975e67bfbd", "Core PE / Energy"),
    ("Learn Capital IV",             "8108570100", "0f53a356-651b-485b-b70f-4a024f6aac0a", "VC / Venture"),
    ("DFJ Growth IV",                "8108570100", "0f53a356-651b-485b-b70f-4a024f6aac0a", "VC / Growth"),
    ("HPS Core Senior Lending II",   "8104716100", "d91d5307-2c1b-46ce-a33b-cdd2ba61be42", "Credit / Direct Lending"),
    ("Apollo EPF IV",                "8104710100", "d91d5307-2c1b-46ce-a33b-cdd2ba61be42", "Credit / Distressed"),
    ("GSO Energy",                   "8104710100", "d91d5307-2c1b-46ce-a33b-cdd2ba61be42", "Credit / Real Asset"),
    ("Sculptor Real Estate Fund V",  "8104703100", "3c20198f-2c5c-433d-a72d-b2de598f381b", "RA / Opp Value-Add"),
    ("ASF IX",                       "8104718100", "364b461b-96cb-4c73-9048-e7f78eb8c407", "RA / Infrastructure"),
    ("Water Property Investor II",   "9006185000", "4464fe55-26c7-4e3b-bfc0-c691901b4590", "RA / Timber"),
]

PUBLIC = [
    ("Microsoft Corp",                         "US5949181045", "US5949181045", 0.10),
    ("JPM US Select Equity Fund",              "LU0672672143", "LU0672672143", 0.10),
    ("ABN Parnassus US Sustainable Equity",    "LU1670606760", "LU1670606760", 0.15),
    ("HSBC Global Aggregate Bond Index",       "IE00BF5DZ473", "IE00BF5DZ473", 0.10),
    ("JPM Aggregate Bond EUR Hedged",          "LU0430493568", "LU0430493568", 0.10),
    ("HSBC Global Aggregate Bond Index",       "IE00BF5DZ473", "IE00BF5DZ473", 0.15),
    ("Vanguard Global Bond Index EUR Hedged",  "IE00BGCZ0B53", "IE00BGCZ0B53", 0.10),
]

wb = Workbook()
ws = wb.active; ws.title = "Portfolio Risk"

HDR = PatternFill('solid', fgColor='1B2A4A')
PI_BG = PatternFill('solid', fgColor='DAEEF3')
TOT_BG = PatternFill('solid', fgColor='2F5496')
hf = Font(name='Arial', size=8, bold=True, color='FFFFFF')
bf = Font(name='Arial', size=8); bfb = Font(name='Arial', size=8, bold=True)
pif = Font(name='Arial', size=8, bold=True, color='2F5496')
tf = Font(name='Arial', size=8, bold=True, color='FFFFFF')
thin = Side(style='thin', color='CCCCCC')
brd = Border(top=thin, bottom=thin, left=thin, right=thin)
ga = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Parameters
ws.cell(row=1, column=2, value="Parameters").font = Font(name='Arial', size=10, bold=True)
for r, (k, v) in enumerate([("Currency","USD"),("Horizon","252d"),("Confidence",0.95),
    ("GFC 2008",3021),("+100bps USD",21148),("-100bps USD",21157)], 2):
    ws.cell(row=r, column=2, value=k).font = bf; ws.cell(row=r, column=3, value=v).font = bfb

ws.column_dimensions['A'].width = 3; ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 8; ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 14; ws.column_dimensions['G'].width = 36; ws.column_dimensions['H'].width = 7

pos_headers = ["Vol","Vol Contrib","ES","ES Contrib","VaR","VaR Contrib","GFC","+100bp","-100bp"]
for c in range(9, 9+len(pos_headers)):
    ws.column_dimensions[get_column_letter(c)].width = 10

cur_row = 9
for pi_idx, (pi_name, pi_jpm, pi_elid, pi_cat) in enumerate(PI_FUNDS):
    r0 = cur_row
    ws.cell(row=r0, column=5, value=f"PTF {pi_idx+1}: {pi_name}").font = Font(name='Arial', size=10, bold=True, color='1B2A4A')
    ws.cell(row=r0, column=13, value=pi_cat).font = Font(name='Arial', size=8, italic=True, color='888888')

    r_hdr = r0 + 1
    for i, h in enumerate(["name","JPM ID","EL ID","Wt"] + pos_headers):
        c = ws.cell(row=r_hdr, column=5+i, value=h)
        c.font = hf; c.fill = HDR; c.alignment = ga; c.border = brd

    r_data = r_hdr + 1
    positions = PUBLIC[:5] + [(pi_name, pi_jpm, pi_elid, 0.20)] + PUBLIC[5:]
    r_end = r_data + len(positions) - 1
    g = f"$G${r_data}:$G${r_end}"; h = f"$H${r_data}:$H${r_end}"

    for i, (name, jpm_id, el_id, weight) in enumerate(positions):
        r = r_data + i; is_pi = (i == 5)
        ws.cell(row=r, column=5, value=name).font = pif if is_pi else bf
        ws.cell(row=r, column=6, value=jpm_id).font = bf
        ws.cell(row=r, column=7, value=el_id).font = bf
        ws.cell(row=r, column=8, value=weight).font = bfb; ws.cell(row=r, column=8).number_format = '0%'
        if is_pi:
            for c in range(5, 18): ws.cell(row=r, column=c).fill = PI_BG
        for c in range(5, 18): ws.cell(row=r, column=c).border = brd

    # Position-level array formulas
    for col, formula in [
        (9,  f'=_xldudf_ELRISKMEASURE("volatility","positions",{g},{h},"weight","USD",252,0.95)'),
        (10, f'=_xldudf_ELRISKMEASURE("volatility","contributions",{g},{h},"weight","USD",252,0.95)'),
        (11, f'=_xldudf_ELRISKMEASURE("es","positions",{g},{h},"weight","USD",252,0.95)'),
        (12, f'=_xldudf_ELRISKMEASURE("es","contributions",{g},{h},"weight","USD",252,0.95)'),
        (13, f'=_xldudf_ELRISKMEASURE("var","positions",{g},{h},"weight","USD",252,0.95)'),
        (14, f'=_xldudf_ELRISKMEASURE("var","contributions",{g},{h},"weight","USD",252,0.95)'),
        (15, f'=_xldudf_ELSTRESSSCENARIO("pnl","positions",3021,{g},{h},"weight","USD","relative")'),
        (16, f'=_xldudf_ELSTRESSSCENARIO("pnl","positions",21148,{g},{h},"weight","USD","relative")'),
        (17, f'=_xldudf_ELSTRESSSCENARIO("pnl","positions",21157,{g},{h},"weight","USD","relative")'),
    ]:
        ref = f"{get_column_letter(col)}{r_data}:{get_column_letter(col)}{r_end}"
        ws.cell(row=r_data, column=col).value = ArrayFormula(ref, formula)

    # Portfolio TOTAL row
    r_tot = r_end + 1
    ws.cell(row=r_tot, column=5, value="PORTFOLIO TOTAL").font = tf
    ws.cell(row=r_tot, column=8, value=1.0).font = tf; ws.cell(row=r_tot, column=8).number_format = '0%'
    for c in range(5, 18):
        ws.cell(row=r_tot, column=c).fill = TOT_BG; ws.cell(row=r_tot, column=c).border = brd

    for col, formula in [
        (9,  f'=_xldudf_ELRISKMEASURE("volatility","portfolio",{g},{h},"weight","USD",252,0.95)'),
        (11, f'=_xldudf_ELRISKMEASURE("es","portfolio",{g},{h},"weight","USD",252,0.95)'),
        (13, f'=_xldudf_ELRISKMEASURE("var","portfolio",{g},{h},"weight","USD",252,0.95)'),
        (15, f'=_xldudf_ELSTRESSSCENARIO("pnl","portfolio",3021,{g},{h},"weight","USD","relative")'),
        (16, f'=_xldudf_ELSTRESSSCENARIO("pnl","portfolio",21148,{g},{h},"weight","USD","relative")'),
        (17, f'=_xldudf_ELSTRESSSCENARIO("pnl","portfolio",21157,{g},{h},"weight","USD","relative")'),
    ]:
        ws.cell(row=r_tot, column=col, value=formula).font = tf

    cur_row = r_tot + 2

# PI Fund Mapping tab
ws2 = wb.create_sheet("PI Fund Mapping")
headers = [("Fund Name",35), ("JPM ID",12), ("Asset Class",18), ("Strategy",18),
           ("Burgiss Index",30), ("Custom Asset EL ID",38), ("Source",28)]
for i,(h,w) in enumerate(headers,1):
    c = ws2.cell(row=1,column=i,value=h); c.font=hf; c.fill=HDR; c.alignment=ga; c.border=brd
    ws2.column_dimensions[chr(64+i)].width = w

for i,(n,j,e,cat) in enumerate(PI_FUNDS,2):
    ac = cat.split('/')[0].strip(); st = cat.split('/')[1].strip() if '/' in cat else ''
    burgiss = {"04cc4f53":"Eq Buyout Expansion USD","0f53a356":"Venture Capital USD",
        "d91d5307":"Debt-Senior-USD","3c20198f":"RE Value-Added USA USD",
        "364b461b":"Infrastructure USD","4464fe55":"Timber USD"}.get(e[:8],'')
    ws2.cell(row=i,column=1,value=n).font=bf; ws2.cell(row=i,column=2,value=j).font=bf
    ws2.cell(row=i,column=3,value=ac).font=bf; ws2.cell(row=i,column=4,value=st).font=bf
    ws2.cell(row=i,column=5,value=burgiss).font=bf; ws2.cell(row=i,column=6,value=e).font=bf
    ws2.cell(row=i,column=7,value="SACH mapping (Pragya Singhi)").font=Font(name='Arial',size=7,color='888888')
    for c in range(1,8): ws2.cell(row=i,column=c).border=brd

wb.save(OUTPUT)
print(f"Saved: {OUTPUT} ({cur_row} rows, 10 portfolios)")
